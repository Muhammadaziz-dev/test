from celery import shared_task
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from io import BytesIO
import uuid
import boto3
from django.core.paginator import Paginator
from django.utils import timezone
from django.conf import settings
from django.db.models import Prefetch
from product.models import Product, ProductImage, ExportTaskLog

_s3_client = None
IMAGE_SIZE = (60, 60)  # Target image size
PAGE_SIZE = 200  # Increased page size for better efficiency
PROGRESS_INTERVAL = 50  # Reduced state update frequency


def get_s3_client():
    global _s3_client
    if _s3_client is None:
        _s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        )
    return _s3_client


def create_thumbnail(image_field):
    """Create optimized thumbnail with transparency handling."""
    try:
        with image_field.open(mode='rb') as f:
            pil_image = PILImage.open(f)
            if pil_image.mode in ('RGBA', 'LA', 'P'):
                background = PILImage.new('RGB', pil_image.size, (255, 255, 255))
                background.paste(pil_image, mask=pil_image.split()[3] if pil_image.mode == 'RGBA' else pil_image)
                pil_image = background

            pil_image.thumbnail(IMAGE_SIZE)
            img_buffer = BytesIO()
            pil_image.save(img_buffer, format='JPEG', optimize=True, quality=85)
            img_buffer.seek(0)
            return img_buffer
    except Exception as e:
        return None


@shared_task(bind=True, autoretry_for=(Exception,), retry_kwargs={'max_retries': 3}, retry_backoff=60)
def export_products_excel(self, store_id, task_id=None, user_id=None):
    if not isinstance(task_id, str) or len(task_id) != 36:
        raise ValueError("Invalid task_id format")

    task_log = ExportTaskLog.objects.filter(task_id=task_id).first()
    if not task_log:
        raise ValueError(f"ExportTaskLog not found for task_id: {task_id}")

    task_log.status = 'PROCESSING'
    task_log.started_at = timezone.now()
    task_log.save(update_fields=['status', 'started_at'])

    try:
        image_prefetch = Prefetch(
            'images',
            queryset=ProductImage.objects.order_by('id')[:1],
            to_attr='first_image'
        )

        base_qs = Product.objects.filter(
            store_id=store_id,
            is_deleted=False
        ).prefetch_related(image_prefetch)

        total = base_qs.count()
        processed = 0
        wb = Workbook()
        ws = wb.active
        ws.title = "Mahsulotlar"
        headers = ["ID", "Rasm", "Nomi", "Narxi", "Zaxira"]
        ws.append(headers)

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.freeze_panes = 'A2'

        paginator = Paginator(base_qs, PAGE_SIZE)

        for page_num in paginator.page_range:
            page = paginator.page(page_num)
            for product in page.object_list:
                row = processed + 2
                price = f"{product.out_price:.2f}".rstrip('0').rstrip('.')
                price = f"{price} {product.currency}" if '.' in price else f"{price}.00 {product.currency}"

                ws.cell(row=row, column=1, value=product.id)
                ws.cell(row=row, column=3, value=product.name)
                ws.cell(row=row, column=4, value=price)
                ws.cell(row=row, column=5, value=product.count)

                if hasattr(product, 'first_image') and product.first_image:
                    img_buffer = create_thumbnail(
                        product.first_image[0].thumbnail or product.first_image[0].image
                    )
                    if img_buffer:
                        try:
                            xl_image = XLImage(img_buffer)
                            xl_image.width, xl_image.height = IMAGE_SIZE
                            ws.add_image(xl_image, f'B{row}')
                            ws.row_dimensions[row].height = 45
                        except:
                            pass

                processed += 1
                if processed % PROGRESS_INTERVAL == 0:
                    self.update_state(
                        state='PROGRESS',
                        meta={'current': processed, 'total': total}
                    )
                    task_log.progress = int((processed / total) * 100)
                    task_log.save(update_fields=['progress'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        s3_key = f"exports/store_{store_id}/products_{uuid.uuid4()}.xlsx"
        s3 = get_s3_client()
        s3.upload_fileobj(
            Fileobj=buffer,
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=s3_key,
            ExtraArgs={
                'ContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'ACL': 'private'
            }
        )

        file_url = s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': s3_key},
            ExpiresIn=7 * 24 * 3600
        )

        task_log.file_url = file_url
        task_log.status = 'SUCCESS'
        task_log.completed_at = timezone.now()
        task_log.progress = 100
        task_log.save()

        return file_url

    except Exception as e:
        task_log.status = 'FAILED'
        task_log.error_message = str(e)[:500]
        task_log.completed_at = timezone.now()
        task_log.save()
        raise
